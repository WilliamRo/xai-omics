import os
import numpy as np
import openpyxl
import pickle

from pro.pro_set import PROSet
from collections import OrderedDict
from tframe.data.base_classes import DataAgent
from roma import console


class PROAgent(DataAgent):

  TFD_FILE_NAME = 'pro.tfd'

  @classmethod
  def load(cls):

    from pro_core import th
    dataset = cls.load_as_tframe_data(th.data_dir)

    ratio = th.ratio_of_dataset.split(':')
    ratio = [int(a) for a in ratio]

    if th.cross_validation:
      size = [dataset.size // th.k_fold for _ in range(th.k_fold - 1)]
      size = size + [dataset.size - sum(size)]

      folds = dataset.split(
        size, names=[f'fold_{i + 1}' for i in range(th.k_fold)],
        over_classes=True)

      train_val_set = combine_dataset(
        [folds[i] for i in range(th.k_fold) if i != th.val_fold_index % th.k_fold])
      train_set, valid_set = train_val_set.split(
        9, 1, names=['TrainSet', 'ValidSet'], over_classes=True)

      test_set = folds[th.val_fold_index % th.k_fold]
      test_set.name = 'TestSet'
    else:
      train_set, valid_set, test_set = dataset.split(
        ratio, names=['TrainSet', 'ValidSet', 'TestSet'], over_classes=True)

    feature_mean = np.mean(train_set.features, axis=0)
    feature_std = np.std(train_set.features, axis=0)
    train_set.features = (train_set.features - feature_mean) / feature_std
    valid_set.features = (valid_set.features - feature_mean) / feature_std
    test_set.features = (test_set.features - feature_mean) / feature_std

    return train_set, valid_set, test_set


  @classmethod
  def load_as_tframe_data(cls, data_dir, *args, **kwargs) -> PROSet:
    file_path = os.path.join(data_dir, cls.TFD_FILE_NAME)
    if os.path.exists(file_path): return PROSet.load(file_path)

    # If .tfd file does not exist, try to convert from raw datas
    console.show_status('Trying to convert raw datas to tframe DataSet ...')
    image_dict = cls.load_as_numpy_arrays(data_dir)

    data_set = PROSet(data_dict=image_dict, name='PROSet', NUM_CLASSES=2,
                      classes=['123', '45'])

    # Show status
    console.show_status('Successfully converted {} samples'.format(
      data_set.size))
    # Save DataSet
    console.show_status('Saving datas set ...')
    data_set.save(file_path)
    console.show_status('Data set saved to {}'.format(file_path))
    # Wrap and return

    return data_set


  @classmethod
  def load_as_numpy_arrays(cls, data_dir) -> OrderedDict:

    print(data_dir)
    image_dict = OrderedDict()

    features_path = os.path.join(data_dir, r'prostate_424_ct_pet_norm.xlsx')
    label_path = os.path.join(data_dir, r'label_424.xlsx')

    # Features
    workbook = openpyxl.load_workbook(features_path)
    sheet = workbook.active
    content = []
    for row in sheet.iter_rows(values_only=True):
      content.append(row)
    content = content[1:]
    features = np.array([c[1:] for c in content])
    pids = np.array([c[0] for c in content], dtype=object)

    # Targets
    workbook = openpyxl.load_workbook(label_path)
    sheet = workbook.active
    content = []
    for row in sheet.iter_rows(values_only=True):
      content.append(row)
    content = content[1:]
    targets = np.array([c[2] for c in content])

    targets = np.eye(2)[targets.flatten().astype(int)].astype(np.uint8)

    image_dict['features'] = features
    image_dict['targets'] = targets
    image_dict['pids'] = pids

    return image_dict


def ratio_to_realnum(ratio: list, total_num: int):
  assert len(ratio) > 1
  parts = [int((r / sum(ratio)) * total_num) for r in ratio[:-1]]
  parts.append(total_num - sum(parts))
  assert sum(parts) == total_num

  return parts


def sorting_key(item):
  return int(item.split('_')[0])


def combine_dataset(input):
  items = list(input[0].data_dict.keys())
  data_dict = {}
  for i in items:
    data_dict[i] = np.concatenate([s.data_dict[i] for s in input], axis=0)

  return PROSet(
    data_dict=data_dict, NUM_CLASSES=input[0].num_classes,
    classes=input[0].properties['classes'])



if __name__ == '__main__':
  agent = PROAgent()
  train_set, val_set, test_set = agent.load()
  print()





