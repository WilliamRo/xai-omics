import os
import numpy as np
import SimpleITK as sitk
import openpyxl
import re

from functools import reduce
from tqdm import tqdm



def catch_features_1(input_excel_path, patient_ids):
  '''
  Get age, weight, height, BMI, tPSA
  '''

  feature_dict = {}
  workbook = openpyxl.load_workbook(input_excel_path)
  sheet = workbook.active

  content = [row for row in sheet.iter_rows(values_only=True)][1:]
  assert len(content) == len(patient_ids)

  for p_index, pid in tqdm(enumerate(patient_ids), desc='Catching features 1'):
    assert int(pid.split('_')[0]) == int(content[p_index][1])
    age = int(re.findall(r'\d+', content[p_index][4])[0])
    height = int(re.findall(r'\d+', content[p_index][5])[0])
    weight = int(re.findall(r'\d+', content[p_index][6])[0])
    tpsa = content[p_index][7]
    tpsa = float(re.findall(r'\d+', tpsa)[0]) if type(tpsa) == str else float(tpsa)

    bmi = weight / ((height / 100)**2)

    feature_dict[pid] = [age, height, weight, bmi, tpsa]

  return feature_dict


def catch_features_2(pet_dir, mask_dir, patient_ids, feature_dict1, pet_name):
  '''
  Get SUV MTV TLG
  '''
  feature_dict = {}

  for pid in tqdm(patient_ids, desc='Catching features 2'):
    pet_path = os.path.join(pet_dir, pid, pet_name)
    mask_path = os.path.join(mask_dir, pid, 'lesion.nii.gz')

    pet_image = sitk.ReadImage(pet_path)
    mask_image = sitk.ReadImage(mask_path)

    pet_array = sitk.GetArrayFromImage(pet_image)
    mask_array = sitk.GetArrayFromImage(mask_image)

    space = pet_image.GetSpacing()
    masked_pet_array = pet_array[mask_array == 1]

    # SUV
    suv_mean = np.mean(masked_pet_array)
    suv_std = np.std(masked_pet_array)
    suv_min = np.min(masked_pet_array)
    suv_max = np.max(masked_pet_array)

    # MTV
    mtv = np.sum(mask_array) * reduce(lambda x, y: x * y, space)
    smtv = mtv / feature_dict1[pid][2]

    # TLG
    tlg = suv_mean * mtv
    stlg = suv_mean * smtv

    feature_dict[pid] = [suv_min, suv_max, suv_mean, suv_std, mtv, smtv, tlg, stlg]

  return feature_dict



if __name__ == '__main__':
  '''
  Extract clinical features
  '''
  # Input
  image_dir = r'E:\xai-omics\data\31-Prostate\mask_and_image'
  clinical_path = r'E:\xai-omics\data\31-Prostate\raw_data\0324_395例.xlsx'
  mask_dir = image_dir
  patient_ids = os.listdir(image_dir)

  # Output
  save_dir = r'E:\xai-omics\data\31-Prostate\results'

  # Parameter
  pet_name = 'pet_aorta.nii.gz'
  save_file_name = r'features_clinic_395_aorta.xlsx'

  # Start
  feature_dict1 = catch_features_1(clinical_path, patient_ids)
  feature_dict2 = catch_features_2(
    image_dir, mask_dir, patient_ids, feature_dict1, pet_name)


  # Save
  workbook = openpyxl.Workbook()
  sheet = workbook.active
  feature_names = ['age', 'height', 'weight', 'bmi', 'tPSA',
                   'SUV_min', 'SUV_max', 'SUV_mean', 'SUV_std',
                   'MTV', 'sMTV', 'TLG', 'sTLG']
  sheet.append(['PIDS'] + list(feature_names))

  for p in patient_ids:
    sheet.append([p] + feature_dict1[p] + feature_dict2[p])

  workbook.save(os.path.join(save_dir, save_file_name))
  print(f'{save_file_name} saved successfully!')
